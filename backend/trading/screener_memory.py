"""
Screener import memory — Phase 4B.14.

CSV import, long-term stock memory, query helpers.
Paper/research only — no LLM calls; does not create intraday tradecards.
"""

from __future__ import annotations

import csv
import json
import os
import re
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from backend.trading.longterm_scoring import score_longterm_stock
from backend.utils.config import DATA_DIR

IST = ZoneInfo('Asia/Kolkata')
STAGE = '4B.14B'

DEFAULT_IMPORTS_FILE = DATA_DIR / 'screener_imports.jsonl'
DEFAULT_STOCK_MEMORY_FILE = DATA_DIR / 'screener_stock_memory.jsonl'
DEFAULT_IMPORTS_DIR = DATA_DIR / 'imports'

_SUPPORTED_IMPORT_SUFFIXES = ('.csv', '.xlsx')

_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    'company_name': ('name', 'company', 'company name', 'stock name'),
    'symbol': ('symbol', 'nse code', 'nse symbol', 'ticker', 'bse code'),
    'market_cap': (
        'market capitalization', 'market cap', 'mcap',
        'mar cap rs.cr.', 'mar cap rs. cr.', 'mar cap rs cr',
    ),
    'pe': ('stock p/e', 'p/e', 'pe', 'pe ratio', 'price to earning'),
    'debt_to_equity': (
        'debt to equity', 'debt to equity', 'debt / equity', 'debt/equity',
        'debt / eq', 'debt / eq %',
    ),
    'roce': ('return on capital employed', 'roce', 'return on capital employed %', 'roce %'),
    'roe': ('return on equity', 'roe', 'return on equity %', 'roe %'),
    'dividend_payout': (
        'dividend payout', 'dividend payout ratio', 'payout ratio', 'payout ratio %',
        'dividend payout %',
    ),
    'sales_growth': ('sales growth', 'sales growth %', 'revenue growth'),
    'profit_growth': ('profit growth', 'profit growth %', 'net profit growth'),
    'free_cashflow': (
        'free cash flow', 'fcf', 'free cash flow crores',
        'fcf prev ann rs.cr.', 'fcf prev ann rs. cr.', 'fcf prev ann rs cr',
    ),
    'promoter_holding': ('promoter holding', 'promoter holding %', 'promoter %', 'promoter'),
    'pledged_percent': (
        'pledged percentage', 'pledged %', 'promoter shares pledged',
        'promoter pledge', 'pledge %', 'promoter pledged %',
    ),
    'fii_holding': ('fii holding', 'fii %', 'fii', 'fii holding %'),
    'dii_holding': ('dii holding', 'dii %', 'dii', 'dii holding %'),
    'mutual_fund_holding': ('mf holding', 'mutual fund', 'mutual fund holding', 'mf %'),
    'public_holding': ('public holding', 'public %', 'public holding %'),
    'retail_holding': ('retail', 'retail holding', 'retail %'),
    'govt_holding': ('government', 'govt holding', 'govt %'),
    'insurance_holding': ('insurance', 'insurance holding', 'insurance %'),
    'number_of_shareholders': (
        'no. of shareholders', 'number of shareholders', 'shareholders', 'no of shareholders',
    ),
    'promoter_holding_change_qoq': (
        'promoter holding change', 'promoter chg qoq', 'promoter holding change qoq',
    ),
    'promoter_pledge_change_qoq': ('promoter pledge change', 'pledge change qoq'),
    'fii_holding_change_qoq': ('fii change', 'fii holding change', 'fii chg qoq'),
    'dii_holding_change_qoq': ('dii change', 'dii holding change', 'dii chg qoq'),
    'current_price': ('current price', 'price', 'cmp', 'close', 'cmp rs.', 'cmp rs'),
    'avg_volume': ('average volume', 'avg volume', 'volume', 'traded volume'),
}


def imports_file_path() -> Path:
    override = os.environ.get('SCREENER_IMPORTS_FILE', '').strip()
    return Path(override) if override else DEFAULT_IMPORTS_FILE


def stock_memory_file_path() -> Path:
    override = os.environ.get('SCREENER_STOCK_MEMORY_FILE', '').strip()
    return Path(override) if override else DEFAULT_STOCK_MEMORY_FILE


def imports_dir_path() -> Path:
    override = os.environ.get('SCREENER_IMPORTS_DIR', '').strip()
    return Path(override) if override else DEFAULT_IMPORTS_DIR


def _now_ist() -> datetime:
    return datetime.now(tz=IST)


def _normalize_header(header: str) -> str:
    return re.sub(r'\s+', ' ', str(header or '').strip().lower())


def _map_headers(fieldnames: list[str] | None) -> dict[str, str]:
    """Map CSV headers to canonical field names."""
    mapping: dict[str, str] = {}
    if not fieldnames:
        return mapping
    normalized = {_normalize_header(h): h for h in fieldnames if h}
    for canonical, aliases in _COLUMN_ALIASES.items():
        for alias in aliases:
            key = _normalize_header(alias)
            if key in normalized:
                mapping[canonical] = normalized[key]
                break
    return mapping


def _normalize_symbol(value: object) -> str:
    text = str(value or '').strip().upper()
    text = re.sub(r'[^A-Z0-9&-]', '', text)
    return text


def strip_screener_query(text: str) -> str:
    raw = str(text or '').strip()
    if len(raw) >= 2 and raw[0] == raw[-1] and raw[0] in ('"', "'"):
        return raw[1:-1].strip()
    return raw


def _normalize_company_match(text: str) -> str:
    return re.sub(r'\s+', ' ', str(text or '').strip().lower())


def _looks_like_nse_symbol(value: object) -> bool:
    text = str(value or '').strip()
    if not text or ' ' in text:
        return False
    norm = _normalize_symbol(text)
    return len(norm) >= 2 and norm == re.sub(r'[^A-Z0-9&-]', '', text.upper())


def _collapse_spaced_acronym(company_name: str) -> str:
    """Map spaced letter groups like 'I R C T C' to IRCTC."""
    parts = re.split(r'\s+', str(company_name or '').strip())
    if len(parts) >= 3 and all(re.fullmatch(r'[A-Za-z]{1,2}', p or '') for p in parts):
        return _normalize_symbol(''.join(parts))
    return ''


KNOWN_COMPANY_SYMBOL_MAP: dict[str, str] = {
    'i r c t c': 'IRCTC',
    'gillette india': 'GILLETTE',
    'tips music': 'TIPS',
    'abbott india': 'ABBOTINDIA',
    'esab india': 'ESABINDIA',
    'bel': 'BEL',
    'bharat electronics': 'BEL',
    'bharat electronics limited': 'BEL',
}


def resolve_canonical_screener_symbol(row: dict[str, Any]) -> tuple[str, str]:
    """
    Resolve canonical NSE-style symbol + clean company name for long-term memory.
    Never returns single-letter symbols derived from company names.
    """
    company = str(row.get('company_name') or row.get('display_name') or '').strip()
    sym = _normalize_symbol(row.get('symbol') or row.get('symbol_key'))
    name_key = _normalize_company_match(company)

    if name_key in KNOWN_COMPANY_SYMBOL_MAP:
        return KNOWN_COMPANY_SYMBOL_MAP[name_key], company

    collapsed = _collapse_spaced_acronym(company)
    if collapsed and len(collapsed) >= 3:
        return collapsed, company

    if sym and len(sym) >= 2:
        return sym, company

    if company:
        derived = _derive_symbol_key(company)
        if len(derived) >= 2:
            return derived, company

    return (sym if len(sym) >= 2 else ''), company


def _derive_symbol_key(company_name: str) -> str:
    name = str(company_name or '').strip()
    collapsed = _collapse_spaced_acronym(name)
    if collapsed:
        return collapsed
    name_key = _normalize_company_match(name)
    if name_key in KNOWN_COMPANY_SYMBOL_MAP:
        return KNOWN_COMPANY_SYMBOL_MAP[name_key]
    words = re.split(r'\s+', name)
    if not words:
        return ''
    first = re.sub(r'[^A-Za-z0-9&]', '', words[0])
    key = _normalize_symbol(first)
    if len(key) >= 2:
        return key
    if len(words) >= 2:
        combo = _normalize_symbol(''.join(w[0] for w in words[:4] if w))
        if len(combo) >= 2:
            return combo
    return ''


def _stock_display_name(company_name: str, symbol_key: str, *, has_real_symbol: bool) -> str:
    name = str(company_name or '').strip()
    if name:
        return name
    return symbol_key


def resolve_screener_query_exact(query: str) -> dict[str, Any] | None:
    """Exact symbol or company match only — no substring fallback."""
    raw = strip_screener_query(query)
    if not raw:
        return None
    rows = _load_jsonl(stock_memory_file_path(), limit=10000)
    rows.sort(key=lambda r: str(r.get('imported_at') or ''), reverse=True)

    key_norm = _normalize_symbol(raw)
    name_norm = _normalize_company_match(raw)

    alias = KNOWN_COMPANY_SYMBOL_MAP.get(name_norm) or KNOWN_COMPANY_SYMBOL_MAP.get(key_norm.lower())
    if alias:
        for row in rows:
            if _normalize_symbol(row.get('symbol_key') or row.get('symbol')) == alias:
                return row

    for row in rows:
        row_key = _normalize_symbol(row.get('symbol_key') or row.get('symbol'))
        if row_key and row_key == key_norm:
            return row
    for row in rows:
        if _normalize_company_match(row.get('company_name')) == name_norm:
            return row
    return None


def resolve_screener_query(query: str) -> dict[str, Any] | None:
    """Resolve Screener memory row by symbol_key, symbol, or company name."""
    raw = strip_screener_query(query)
    if not raw:
        return None
    if _looks_like_nse_symbol(raw):
        exact = resolve_screener_query_exact(raw)
        if exact:
            return exact
        return None
    rows = _load_jsonl(stock_memory_file_path(), limit=10000)
    rows.sort(key=lambda r: str(r.get('imported_at') or ''), reverse=True)

    key_norm = _normalize_symbol(raw)
    name_norm = _normalize_company_match(raw)

    for row in rows:
        row_key = _normalize_symbol(row.get('symbol_key') or row.get('symbol'))
        if row_key and row_key == key_norm:
            return row
    for row in rows:
        if _normalize_company_match(row.get('company_name')) == name_norm:
            return row
    for row in rows:
        cn = _normalize_company_match(row.get('company_name'))
        dn = _normalize_company_match(row.get('display_name'))
        if cn and (cn.startswith(name_norm) or name_norm in cn):
            return row
        if dn and (dn.startswith(name_norm) or name_norm in dn):
            return row
    return None


def _row_value(raw: dict[str, Any], header_map: dict[str, str], field: str) -> Any:
    src = header_map.get(field)
    if not src:
        return None
    val = raw.get(src)
    if val in (None, '', '-', 'NA', 'N/A'):
        return None
    return val


def _parse_csv_rows(path: Path) -> tuple[list[dict[str, Any]], list[str], dict[str, str]]:
    with path.open('r', encoding='utf-8-sig', newline='') as handle:
        reader = csv.DictReader(handle)
        fieldnames = list(reader.fieldnames or [])
        header_map = _map_headers(fieldnames)
        rows = [dict(r) for r in reader if _row_has_data(r)]
    return rows, fieldnames, header_map


def _cell_text(value: object) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _row_has_data(raw: dict[str, Any]) -> bool:
    return any(_cell_text(v) for v in raw.values())


def _is_repeated_header_row(cells: tuple[object, ...], header_norms: set[str]) -> bool:
    if not header_norms:
        return False
    norms = {_normalize_header(_cell_text(c)) for c in cells if _cell_text(c)}
    if len(norms) < 2:
        return False
    overlap = norms & header_norms
    return len(overlap) >= max(2, len(norms) // 2)


def _parse_xlsx_rows(path: Path) -> tuple[list[dict[str, Any]], list[str], dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError('openpyxl is required for XLSX import') from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        headers: list[str] = []
        header_map: dict[str, str] = {}
        header_norms: set[str] = set()
        rows: list[dict[str, Any]] = []
        for row in ws.iter_rows(values_only=True):
            if not row or all(_cell_text(c) == '' for c in row):
                continue
            if not headers:
                headers = [_cell_text(c) for c in row]
                header_map = _map_headers(headers)
                header_norms = {_normalize_header(h) for h in headers if h}
                continue
            if _is_repeated_header_row(row, header_norms):
                continue
            row_dict = {
                headers[i]: row[i] if i < len(row) else None
                for i in range(len(headers))
            }
            if _row_has_data(row_dict):
                rows.append(row_dict)
    finally:
        wb.close()
    return rows, headers, header_map


def parse_screener_rows(path: Path) -> tuple[list[dict[str, Any]], list[str], dict[str, str], str]:
    suffix = path.suffix.lower()
    if suffix == '.csv':
        rows, fieldnames, header_map = _parse_csv_rows(path)
        return rows, fieldnames, header_map, 'screener_csv'
    if suffix == '.xlsx':
        rows, fieldnames, header_map = _parse_xlsx_rows(path)
        return rows, fieldnames, header_map, 'screener_xlsx'
    raise ValueError(f'Unsupported file type: {suffix or path.name}')


def parse_screener_import_filename(args: str) -> str:
    """
    Extract filename from screener import args.

    '/screener import longterm file.xlsx' -> args 'import longterm file.xlsx' -> 'file.xlsx'
    """
    raw = str(args or '').strip()
    lower = raw.lower()
    if lower in ('', 'import longterm', 'longterm'):
        return ''
    marker = 'import longterm '
    if lower.startswith(marker):
        return raw[len(marker):].strip()
    if lower.startswith('longterm '):
        return raw.split(None, 1)[1].strip()
    return ''


def sanitize_import_filename(name: str) -> str:
    base = Path(str(name or '').strip()).name
    base = re.sub(r'[^A-Za-z0-9._-]+', '_', base).strip('._')
    if not base:
        raise ValueError('invalid filename')
    lower = base.lower()
    if not lower.endswith(_SUPPORTED_IMPORT_SUFFIXES):
        raise ValueError('filename must end with .csv or .xlsx')
    return base


def _normalize_stock_row(
    raw: dict[str, Any],
    header_map: dict[str, str],
    *,
    import_id: str,
    imported_at: str,
    screen_name: str,
) -> dict[str, Any] | None:
    company_name = str(_row_value(raw, header_map, 'company_name') or '').strip()
    symbol_raw = _row_value(raw, header_map, 'symbol')
    has_real_symbol = bool(symbol_raw) and _looks_like_nse_symbol(symbol_raw)

    if has_real_symbol:
        symbol_key = _normalize_symbol(symbol_raw)
    elif company_name:
        symbol_key = _derive_symbol_key(company_name)
    else:
        return None

    if not symbol_key:
        return None

    display_name = _stock_display_name(company_name, symbol_key, has_real_symbol=has_real_symbol)

    base = {
        'import_id': import_id,
        'imported_at': imported_at,
        'symbol': symbol_key,
        'symbol_key': symbol_key,
        'company_name': display_name,
        'display_name': display_name,
        'market_cap': _row_value(raw, header_map, 'market_cap'),
        'pe': _row_value(raw, header_map, 'pe'),
        'debt_to_equity': _row_value(raw, header_map, 'debt_to_equity'),
        'roce': _row_value(raw, header_map, 'roce'),
        'roe': _row_value(raw, header_map, 'roe'),
        'dividend_payout': _row_value(raw, header_map, 'dividend_payout'),
        'sales_growth': _row_value(raw, header_map, 'sales_growth'),
        'profit_growth': _row_value(raw, header_map, 'profit_growth'),
        'free_cashflow': _row_value(raw, header_map, 'free_cashflow'),
        'promoter_holding': _row_value(raw, header_map, 'promoter_holding'),
        'pledged_percent': _row_value(raw, header_map, 'pledged_percent'),
        'fii_holding': _row_value(raw, header_map, 'fii_holding'),
        'dii_holding': _row_value(raw, header_map, 'dii_holding'),
        'mutual_fund_holding': _row_value(raw, header_map, 'mutual_fund_holding'),
        'public_holding': _row_value(raw, header_map, 'public_holding'),
        'retail_holding': _row_value(raw, header_map, 'retail_holding'),
        'govt_holding': _row_value(raw, header_map, 'govt_holding'),
        'insurance_holding': _row_value(raw, header_map, 'insurance_holding'),
        'number_of_shareholders': _row_value(raw, header_map, 'number_of_shareholders'),
        'promoter_holding_change_qoq': _row_value(raw, header_map, 'promoter_holding_change_qoq'),
        'promoter_pledge_change_qoq': _row_value(raw, header_map, 'promoter_pledge_change_qoq'),
        'fii_holding_change_qoq': _row_value(raw, header_map, 'fii_holding_change_qoq'),
        'dii_holding_change_qoq': _row_value(raw, header_map, 'dii_holding_change_qoq'),
        'current_price': _row_value(raw, header_map, 'current_price'),
        'avg_volume': _row_value(raw, header_map, 'avg_volume'),
        'screen_name': screen_name,
    }
    scored = score_longterm_stock(base)
    base.update(scored)
    return base


def _append_jsonl(path: Path, record: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('a', encoding='utf-8') as handle:
        handle.write(json.dumps(record, ensure_ascii=False) + '\n')


def _load_jsonl(path: Path, limit: int = 10000) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows: list[dict[str, Any]] = []
    try:
        with path.open('r', encoding='utf-8') as handle:
            for line in handle:
                line = line.strip()
                if not line:
                    continue
                try:
                    row = json.loads(line)
                except json.JSONDecodeError:
                    continue
                if isinstance(row, dict):
                    rows.append(row)
    except OSError:
        return []
    return rows[-limit:]


def append_screener_import(record: dict[str, Any]) -> dict[str, Any]:
    payload = dict(record)
    payload.setdefault('import_id', uuid.uuid4().hex)
    _append_jsonl(imports_file_path(), payload)
    return payload


def append_stock_memory(record: dict[str, Any]) -> dict[str, Any]:
    _append_jsonl(stock_memory_file_path(), dict(record))
    return record


def load_screener_imports(limit: int = 50) -> list[dict[str, Any]]:
    rows = _load_jsonl(imports_file_path(), limit=limit)
    rows.sort(key=lambda r: str(r.get('imported_at') or ''), reverse=True)
    return rows[:limit]


def load_stock_memory(symbol: str | None = None, limit: int = 500) -> list[dict[str, Any]]:
    rows = _load_jsonl(stock_memory_file_path(), limit=10000)
    rows.sort(key=lambda r: str(r.get('imported_at') or ''), reverse=True)
    if symbol:
        match = resolve_screener_query(symbol)
        if match:
            key = _normalize_symbol(match.get('symbol_key') or match.get('symbol'))
            rows = [
                r for r in rows
                if _normalize_symbol(r.get('symbol_key') or r.get('symbol')) == key
            ]
        else:
            sym = _normalize_symbol(strip_screener_query(symbol))
            rows = [
                r for r in rows
                if _normalize_symbol(r.get('symbol_key') or r.get('symbol')) == sym
            ]
    return rows[:limit]


def latest_import() -> dict[str, Any] | None:
    imports = load_screener_imports(limit=1)
    return imports[0] if imports else None


def latest_import_stocks(limit: int = 500) -> list[dict[str, Any]]:
    imp = latest_import()
    if not imp:
        return []
    import_id = str(imp.get('import_id') or '')
    rows = load_stock_memory(limit=5000)
    matched = [r for r in rows if str(r.get('import_id') or '') == import_id]
    matched.sort(key=lambda r: int(r.get('longterm_score') or 0), reverse=True)
    return matched[:limit]


def screener_status() -> dict[str, Any]:
    imp = latest_import()
    all_stocks = _load_jsonl(stock_memory_file_path(), limit=100000)
    symbols = {_normalize_symbol(r.get('symbol')) for r in all_stocks if _normalize_symbol(r.get('symbol'))}
    return {
        'latest_import': imp,
        'total_stocks': len(symbols),
        'total_records': len(all_stocks),
    }


def summarize_symbol_screener(query: str) -> dict[str, Any]:
    raw = strip_screener_query(query)
    sym = _normalize_symbol(raw)
    latest = resolve_screener_query(raw)
    if not latest:
        return {'symbol': sym, 'symbol_key': sym, 'count': 0}
    sym_key = _normalize_symbol(latest.get('symbol_key') or latest.get('symbol'))
    rows = load_stock_memory(sym_key, limit=50)
    display_name = str(latest.get('display_name') or latest.get('company_name') or sym_key)
    return {
        'symbol': sym_key,
        'symbol_key': sym_key,
        'company_name': str(latest.get('company_name') or display_name),
        'display_name': display_name,
        'count': len(rows),
        'latest': latest,
        'imported_at': str(latest.get('imported_at') or ''),
        'screen_name': str(latest.get('screen_name') or ''),
        'longterm_score': int(latest.get('longterm_score') or 0),
        'verdict': str(latest.get('verdict') or 'unknown'),
        'reasons': list(latest.get('reasons') or [])[:5],
        'risk_flags': list(latest.get('risk_flags') or [])[:5],
        'cap_bucket': str(latest.get('cap_bucket') or 'unknown'),
    }


def import_screener_file(
    filepath: Path | str,
    *,
    screen_name: str = '',
    query_text: str = '',
    notes: str = '',
) -> dict[str, Any]:
    """
    Import Screener CSV or XLSX into memory stores.

    Does not create intraday tradecards.
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f'Screener file not found: {path}')

    raw_rows, _fieldnames, header_map, source = parse_screener_rows(path)
    ist_now = _now_ist()
    import_id = uuid.uuid4().hex
    imported_at = ist_now.replace(microsecond=0).isoformat()
    screen = str(screen_name or path.stem).strip() or 'screener_import'
    query = str(query_text or screen).strip()

    import_record = {
        'import_id': import_id,
        'imported_at': imported_at,
        'current_ist': ist_now.strftime('%Y-%m-%d %H:%M IST'),
        'screen_name': screen,
        'query_text': query,
        'source': source,
        'filename': path.name,
        'row_count': len(raw_rows),
        'normalized_columns': sorted(header_map.keys()),
        'notes': notes,
    }
    append_screener_import(import_record)

    stored: list[dict[str, Any]] = []
    for raw in raw_rows:
        stock = _normalize_stock_row(
            raw,
            header_map,
            import_id=import_id,
            imported_at=imported_at,
            screen_name=screen,
        )
        if stock:
            append_stock_memory(stock)
            stored.append(stock)

    result = {
        'ok': True,
        'import': import_record,
        'stored_count': len(stored),
        'stored_stocks': stored,
        'skipped': max(0, len(raw_rows) - len(stored)),
    }
    try:
        from backend.trading.longterm_snapshot_memory import capture_screener_import_snapshot

        capture_screener_import_snapshot(result, source_file_name=path.name)
    except Exception:
        pass
    try:
        from backend.trading.weekly_signal_capture import capture_screener_import_signals

        capture_screener_import_signals(stored)
    except Exception:
        pass
    try:
        from backend.trading.investor_intelligence import capture_investor_from_screener_stocks

        capture_investor_from_screener_stocks(
            stored,
            import_id=import_id,
            imported_at=imported_at,
        )
    except Exception:
        pass
    return result


def import_screener_csv(
    filepath: Path | str,
    *,
    screen_name: str = '',
    query_text: str = '',
    notes: str = '',
) -> dict[str, Any]:
    """Backward-compatible alias for import_screener_file."""
    return import_screener_file(
        filepath,
        screen_name=screen_name,
        query_text=query_text,
        notes=notes,
    )


def resolve_import_filepath(filename: str) -> Path:
    """Resolve CSV/XLSX path under imports dir."""
    raw = Path(str(filename or '').strip()).name
    if not raw:
        raise ValueError('filename required')
    if not raw.lower().endswith(_SUPPORTED_IMPORT_SUFFIXES):
        raw = f'{raw}.csv'
    name = sanitize_import_filename(raw)
    return imports_dir_path() / name


def save_import_bytes(data: bytes, filename: str) -> Path:
    """Save uploaded bytes to imports dir with sanitized filename."""
    safe_name = sanitize_import_filename(filename)
    dest = imports_dir_path() / safe_name
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(data)
    return dest
